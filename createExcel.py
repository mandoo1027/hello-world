from openpyxl import Workbook # 엑셀 라이브러리 가져오기
from openpyxl.styles import Font,Alignment ,Border, Side
import datetime
import cx_Oracle              # 오라클 라이브러리 가져오기

##############################DB접속###########################
#오라클 접속
port_num = 1521
dsn = cx_Oracle.makedsn("192.168.63.213",port_num,"ORCL")
con = cx_Oracle.connect("HR52_DEV","hr52_dev12##",dsn)
cursor = con.cursor()
##############################DB접속###########################
############################파일명 선언########################
#생성될 파일이름
saveFileNam= "test.xlsx"
#긁어올 헤더정보
selectHeaderFileName="./sql/getHeader.txt"
#데이터정보 불러오기(쿼리)
selectDataSql="./sql/selectHeader.txt"
############################파일명 선언########################


#엑셀 컴포넌트
wb = Workbook() #excel 객체 가져오기
ws1 = wb.active #현재 열려 있는 Sheet
ws1.title = "일근무리스트" #시트 ws1의 이름을 지정 
ws1.merge_cells('A1:X2') #셀 합치기

#엑셀 제목 셋팅
ws1['A1'] = '일근무리스트 조회' # 병합된 셀에 제목 셋팅

ws1['A3'] = "작성자"  
ws1['B3'] = "아무개"  #작성자
ws1['C3'] = "작성일자"
ws1['D3'] = datetime.datetime.today().strftime("%Y/%m/%d %H:%M:%S") # 생성일자
A1 = ws1['A1']
A1.font = Font(name="맑은 고딕", size = 15 , bold = True) #글자 셋팅
A1.alignment = Alignment(horizontal="center",vertical='center') #가운데 정렬 

ws1.freeze_panes = 'B5' #엑셀 고정

headerCnt = 0 #헤더 갯수



#헤더 시작 로우
headerRow = 4
#데이터 시작 로우
dataStartRow = 5


#파일 내용 Text로 가져오기
def getFileText(fileName): 
	sqlStr = "";
	f = open(fileName,'r')
	sqlStr = f.read() #파일 내용 가져오기 
	f.close()
	return sqlStr
#셀별 가운데 정렬 
def setCenter(objectvlaue):
		objectvlaue = Alignment(horizontal="center",vertical='center') 

#쿼리로 테이블에 컬럼 코멘드를 조회하여 엑셀 상단 타이틀 셋팅
def excelHeaderSet():
	selectColumn = getFileText(selectDataSql)#해당 파일의 쿼리 긁어오기 
	cursor.execute(selectColumn)
	rows = cursor.fetchall() #모든 정보 가져오기  한줄씩 가져오는 명령어 cursor.fetchone()->  이땐 while문 이용
	headerCnt = len(rows) # 헤더 갯수 셋팅
	for col in range(1,len(rows)):
		#row = 3    row1~2까지는 제목이기때문에 3부터 시작
		ws1.cell(row=headerRow,column=col,value="{}".format(rows[col][2]))
		ws1.cell(row=headerRow,column=col).font = Font( name='Calibri', size=12, bold=True, italic=False, color='FF000000' )#헤더 색깔 변경
		setCenter(ws1.cell(row=headerRow,column=col).alignment)#가운데 정렬
		
#텍스트를 조회하여 엑셀 항목을 생성하는 함수
def getTextHeaderSet():
	headerList = []
	file = open(selectHeaderFileName,'r')
	for str in file.readlines():
		headerList.append(str.strip("\n"))
	file.close() 
	
	for col in range(1,len(headerList)):
		#row = 4    row1~2까지는 제목이기때문에 4부터 시작
		ws1.cell(row=headerRow,column=col,value="{}".format(headerList[col-1]))#값 셋팅
		ws1.cell(row=headerRow,column=col).font = Font( name='Calibri', size=12, bold=True, italic=False, color='FF000000' )#헤더 색깔 변경
		setCenter(ws1.cell(row=headerRow,column=col).alignment)#가운데 정렬		
		 
		
#엑셀 데이터 리스트 셋팅
def excelDataSet():
	selectColumn = getFileText("./sql/selectData.txt")#해당 파일의 쿼리 긁어오기 
	cursor.execute(selectColumn)
	rows = cursor.fetchall() #모든 정보 가져오기  한줄씩 가져오는 명령어 cursor.fetchone()->  이땐 while문 이용
	colCnt = 0
	if(len(rows) > 0):
		colCnt = len(rows[0]) #컬럼 카운트
	
	excelRow = dataStartRow#제목,헤더포함 4부터 시작
	for rowNum in range(0,len(rows)): 
		for col in range(0,colCnt):
			ws1.cell(row=excelRow, column=col+1, value="{}".format(rows[rowNum][col]))#데이터 셀에 셋팅
			setCenter(ws1.cell(row=excelRow, column=col+1).alignment) #가운데 정렬 
		excelRow = excelRow+1



#엑셀 타이틀 조회(쿼리로 되어있는 정보 긁어오기)
#excelHeaderSet()
#엑셀 타이틀 조회(텍스트로 되어있는 정보 긁어오기)
getTextHeaderSet()
#엑셀 데이터 조회 함수 호출
excelDataSet()

	
cursor.close()	
con.close()

wb.save(saveFileNam)#파일이름 설정